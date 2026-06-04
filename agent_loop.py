import asyncio
import os
import re
import json
from urllib.parse import urlparse

from dotenv import load_dotenv
import requests
from ddgs import DDGS
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig

load_dotenv()

LLM_API = os.getenv("LLM_API_URL", "http://10.0.99.116:8070/v1/chat/completions")
BLOCKED_DOMAINS = [
    "wikipedia", "facebook", "youtube", "twitter", "linkedin",
    "crunchbase", "bloomberg", "reuters", "forbes", "glassdoor",
    "indeed", "zoominfo", "linkedin", "instagram", "tiktok",
]
SPAM_DOMAINS = ["xnxx", "xhamster", "xvideo", "pornhat", "sexvid", "porn", "xxx", "xvideos"]


# ─── LLM ──────────────────────────────────────────────────────────────────────

def call_llm(system: str, user: str) -> str | None:
    payload = {
        "model": "", "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ], "stream": False,
        "options": {"temperature": 0.0, "top_p": 0.95, "top_k": 64, "num_ctx": 40000},
        "chat_template_kwargs": {"enable_thinking": False},
    }
    resp = requests.post(LLM_API, json=payload, stream=False, timeout=120)
    if resp.status_code != 200:
        return None
    return resp.json()["choices"][0]["message"]["content"]


# ─── SEARCH ───────────────────────────────────────────────────────────────────

def search_web(queries: list[str]) -> str:
    if isinstance(queries, str):
        queries = [queries]
    text = ""
    for q in queries:
        with DDGS() as ddgs:
            for r in ddgs.text(q, max_results=5):
                link = r.get("href") or ""
                if any(d in link.lower() for d in SPAM_DOMAINS):
                    continue
                text += f"""Title: {r.get("title", "")}
Link: {link}
Snippet: {r.get("body", "")}
\n"""
    return text


# ─── CRAWL ────────────────────────────────────────────────────────────────────

async def extract_web(url: str) -> str:
    try:
        config = CrawlerRunConfig(magic=True, simulate_user=True, override_navigator=True, page_timeout=15000)
        async with AsyncWebCrawler() as crawler:
            result = await crawler.arun(url=url, config=config)
            md = (result.markdown or "").strip()
            if md:
                return md
    except Exception:
        pass

    try:
        resp = requests.get(url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        })
        if resp.status_code == 200:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(resp.text, "lxml")
            for tag in soup(["script", "style", "nav", "footer", "header"]):
                tag.decompose()
            lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if len(l.strip()) > 30]
            return "\n".join(lines[:200])
    except Exception:
        pass
    return ""


# ─── FIND HOMEPAGE ────────────────────────────────────────────────────────────

def find_homepage(company: str) -> str:
    queries = [f"{company} official website", f"{company} trang chủ"]
    results = []
    for q in queries:
        with DDGS() as ddgs:
            for r in ddgs.text(q, max_results=8):
                link = r.get("href") or ""
                if any(d in link.lower() for d in SPAM_DOMAINS + BLOCKED_DOMAINS):
                    continue
                results.append({"link": link, "title": r.get("title", "")})
    name_parts = company.lower().split()

    def domain(link):
        return urlparse(link).hostname or ""

    def name_in_domain(link):
        d = domain(link).replace("www.", "")
        segs = d.split(".")
        joined = company.lower().replace(" ", "")
        if any(joined == s for s in segs):
            return True
        for p in name_parts:
            if len(p) > 3 and any(p == s for s in segs):
                return True
        return False

    for r in results:
        link = r["link"]
        if name_in_domain(link):
            return link
    for r in results:
        link = r["link"]
        title = r["title"].lower()
        has_c = any(p in title for p in name_parts if len(p) > 3)
        has_o = any(w in title for w in ["official", "home", "trang ch"])
        if has_c and has_o:
            return link
    return results[0]["link"] if results else ""


# ─── MATCH COMPANY ────────────────────────────────────────────────────────────

def match_company(content: str, query: str) -> dict:
    prompt = f"""Bạn là chuyên gia phân tích công ty.

Dưới đây là nội dung trang chủ:
<homepage>
{content[:8000]}
</homepage>

Nhiệm vụ:
1. Xác định công ty có cung cấp sản phẩm/giải pháp liên quan đến "{query}" không
2. Nếu có, liệt kê sản phẩm liên quan

Chỉ trả về JSON:
{{"answer": true/false, "products": ["sp1", "sp2"]}}"""
    resp = call_llm("Bạn là chuyên gia phân tích.", prompt)
    if not resp:
        return {"answer": False, "products": []}
    try:
        m = re.search(r"\{.*\}", resp, re.DOTALL)
        return json.loads(m.group(0)) if m else {"answer": False, "products": []}
    except Exception:
        return {"answer": False, "products": []}


# ─── EXTRACT JSON ────────────────────────────────────────────────────────────

def extract_json(text: str):
    m = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL)
    if m:
        text = m.group(1)
    return json.loads(text.strip())


# ─── EXCEL ────────────────────────────────────────────────────────────────────

def read_priority_exhibitions(filepath: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb["AIPT Priority Shortlist"]
    lines = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 5).value
        if not name:
            continue
        lines.append(f"{name} | Website: {ws.cell(row, 6).value or ''} | Category: {ws.cell(row, 7).value or ''} | Tags: {ws.cell(row, 8).value or ''}")
    return "\n".join(lines)


# ─── AGENT LOOP ──────────────────────────────────────────────────────────────

async def main():
    # ── Config ──
    excel_path = r"D:\Search_Imex\AIPT_Global_Defense_Security_Exhibitions_Full_Database.xlsx"
    query = "Máy bay không người lái"
    MAX_STEPS = 20

    # ── Init ──
    exhibitions = read_priority_exhibitions(excel_path)
    context = ""
    history = ""
    visited_urls = set()
    found_companies: list[str] = []

    system_prompt = """Bạn là chuyên gia phân tích web tìm hãng cung cấp sản phẩm.

<tools>
search_web: Tìm kiếm thông tin trên web. Input là search query string.
extract_web: Trích xuất nội dung từ URL. Input là một URL duy nhất.
answer: Kết thúc, trả về danh sách hãng.
</tools>

<rules>
- Mỗi lần chỉ gọi MỘT tool duy nhất.
- Không gọi lại tool với input đã dùng.
- Không visit URL đã có trong visited_urls.
- Nội dung trả về phải là JSON array, mỗi phần tử:
  {"reason": "lý do", "next_action": "search_web|extract_web|answer", "input": "...", "context": "tóm tắt", "imex": ["hãng1"]}
- Khi có đủ thông tin, trả về answer kèm danh sách hãng.
- Nếu đạt bước tối đa, trả về answer.
</rules>"""

    # ── Loop ──
    for step in range(MAX_STEPS):
        if len(found_companies) >= 5:
            break

        print(f"\n===== STEP {step+1}/{MAX_STEPS} =====")

        user = f"""<query>{query}</query>
<step>{step+1}/{MAX_STEPS}</step>
<visited_urls>{", ".join(visited_urls) or "chưa có"}</visited_urls>
<history>{history[-5000:]}</history>
<context>{context[-40000:]}</context>
<exhibitions>{exhibitions}</exhibitions>
<task>Tìm hãng cung cấp sản phẩm "{query}". Quyết định 1 hành động duy nhất.</task>"""

        resp = call_llm(system_prompt, user)
        if not resp:
            print("LLM không phản hồi")
            continue

        print(f"LLM: {resp[:200]}")

        try:
            actions = extract_json(resp)
        except Exception as e:
            print(f"Parse JSON lỗi: {e}")
            continue

        if isinstance(actions, dict):
            actions = [actions]

        for act in actions:
            action = act.get("next_action", "")
            inp = act.get("input", "")

            if action == "answer":
                imex = act.get("imex", [])
                if isinstance(imex, list):
                    for c in imex:
                        if c and c not in found_companies:
                            found_companies.append(c)
                break

            print(f"→ {action}: {inp[:100]}")

            if action == "search_web":
                data = search_web(inp)
                context += f"\n--- Search: {inp} ---\n{data[:5000]}"
                history += f"[S{step+1}] search_web('{inp}')\n"

            elif action == "extract_web":
                if inp in visited_urls:
                    print(f"  Bỏ qua URL đã visit: {inp}")
                    continue
                visited_urls.add(inp)
                data = await extract_web(inp)
                if data:
                    context += f"\n--- Extract: {inp} ---\n{data[:10000]}"
                history += f"[S{step+1}] extract_web('{inp}')\n"

            imex = act.get("imex", [])
            if isinstance(imex, list):
                for c in imex:
                    if c and c not in found_companies:
                        found_companies.append(c)
                        print(f"  + Hãng: {c}")

            if len(found_companies) >= 5:
                break

    print(f"\n{'='*50}")
    print(f"Tìm thấy {len(found_companies)} hãng. Đang verify...")

    # ── Verify ──
    verified = []
    for company in found_companies:
        if len(verified) >= 5:
            break
        print(f"\nVerify: {company}")
        hp = find_homepage(company)
        if not hp:
            print(f"  ❌ Không tìm thấy homepage")
            verified.append({"company": company, "homepage": "", "products": [], "match": False})
            continue
        print(f"  Homepage: {hp}")
        content = await extract_web(hp)
        if content:
            match = match_company(content, query)
            if match.get("answer"):
                print(f"  ✅ Có sản phẩm liên quan: {match.get('products', [])}")
                verified.append({"company": company, "homepage": hp, "products": match.get("products", []), "match": True})
                continue
        print(f"  ❌ Không xác nhận được")
        verified.append({"company": company, "homepage": hp, "products": [], "match": False})

    # ── Report ──
    print(f"\n{'='*50}")
    print("KẾT QUẢ:")
    print("=" * 50)
    for v in verified:
        status = "✅" if v["match"] else "❌"
        sp = ", ".join(v["products"])
        print(f'{status} {v["company"]:30s} | {v["homepage"]:40s} | SP: {sp}')


if __name__ == "__main__":
    asyncio.run(main())
