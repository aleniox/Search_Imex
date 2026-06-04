import asyncio
from crawl4ai import *
import os
from dotenv import load_dotenv
import requests
from ddgs import DDGS
import openpyxl
import re
import json
from urllib.parse import urlparse
load_dotenv()

BLOCKED_DOMAINS = [
    "wikipedia", "facebook", "youtube", "twitter", "linkedin",
    "crunchbase", "bloomberg", "reuters", "forbes", "glassdoor",
    "indeed", "zoominfo", "linkedin", "instagram", "tiktok",
]

SPAM_DOMAINS = ["xnxx", "xhamster", "xvideo", "pornhat", "sexvid", "porn", "xxx", "xvideos"]
LLM_API = os.getenv("LLM_API_URL", "http://10.0.99.116:8070/v1/chat/completions")


def call_chat_api(messages, model="", stream=True, max_token=40000, host=LLM_API):
    payload = {
        "model": model, "messages": messages, "stream": stream,
        "options": {"temperature": 0.0, "top_p": 0.95, "top_k": 64, "num_ctx": max_token},
        "chat_template_kwargs": {"enable_thinking": False}
    }
    return requests.post(host, json=payload, stream=stream)

def search_web(queries: list[str], return_text: bool=True) -> list[dict]:
        all_results = []
        all_text = ""
        if isinstance(queries,str):
            queries = [queries]
        for query in queries:
            results = []

            with DDGS() as ddgs:
                for r in ddgs.text(query, max_results=5):
                    link = r.get("href") or ""
                    if any(d in link.lower() for d in SPAM_DOMAINS):
                        continue
                    all_text += f"""
Title: {r.get("title", "")}
Link: {link},
Snippet: {r.get("body", "")}
"""
                    results.append({
                        "title": r.get("title", ""),
                        "link": link,
                        "snippet": r.get("body", ""),
                    })
            all_results.append({"query": query, "results": results})
        if return_text:
            return all_text
        return all_results


def extract_json(text: str):
    # tìm block ```json ... ```
    match = re.search(
        r"```(?:json)?\s*(.*?)\s*```",
        text,
        re.DOTALL
    )

    if match:
        json_text = match.group(1)
    else:
        json_text = text.strip()

    return json.loads(json_text)

def extract_web(url):
    async def _run():
        async with AsyncWebCrawler() as crawler:
            result = await crawler.arun(url=url)
            return result.markdown

    return asyncio.run(_run())

def read_priority_exhibitions(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["AIPT Priority Shortlist"]
    exhibitions = ""
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 5).value
        if not name:
            continue
        exhibitions += f"""Exhibition / Conference: {str(ws.cell(row, 5).value or "")}, 
        Official Website: {str(ws.cell(row, 6).value or "")}, 
        Main Category: {str(ws.cell(row, 7).value or "")}
        Focus Tags: {str(ws.cell(row, 8).value or "")}\n\n"""
    return exhibitions


def find_homepage(company: str) -> str:
    queries = [f"{company} official website", f"{company} trang chủ"]
    results = []
    raw = search_web(queries, return_text=False)
    for r in raw:
        results.extend(r.get("results", []))
    name_parts = company.lower().split()

    def domain(link: str) -> str:
        return urlparse(link).hostname or ""

    def is_blocked(link: str) -> bool:
        d = domain(link)
        return any(b in d for b in BLOCKED_DOMAINS)

    def name_in_domain(link: str) -> bool:
        d = domain(link).replace("www.", "")
        segments = d.split(".")
        joined = company.lower().replace(" ", "")
        if any(joined == seg for seg in segments):
            return True
        for p in name_parts:
            if len(p) > 3 and any(p == seg for seg in segments):
                return True
        return False

    for r in results:
        link = r.get("link", "")
        if name_in_domain(link) and not is_blocked(link):
            return link

    for r in results:
        link = r.get("link", "")
        if is_blocked(link):
            continue
        title = r.get("title", "").lower()
        has_company = any(p in title for p in name_parts if len(p) > 3)
        has_official = any(w in title for w in ["official", "home", "trang ch"])
        if has_company and has_official:
            return link

    for r in results:
        link = r.get("link", "")
        if is_blocked(link):
            continue
        if any(p in link.lower() for p in name_parts if len(p) > 3):
            return link

    for r in results:
        link = r.get("link", "")
        if not is_blocked(link):
            return link

    return ""

tools = {
    "search_web": search_web,
    "extract_web": extract_web
}
# In Jupyter notebooks, just await directly
# data = asyncio.run(main())
data = read_priority_exhibitions(r"D:\Search_Imex\AIPT_Global_Defense_Security_Exhibitions_Full_Database.xlsx")
print(data)
MAX_STEPS = 20
query = "Máy bay không người lái"
system_prompt = """Bạn là một chuyên gia phân tích web để lấy đc các thông tin quan trọng để giúp tìm các hãng liên quan đến sản phẩm do người dùng yêu cầu.
<tools>
search_web: Tìm kiếm thông tin trên web (input: str) Input MUST be a search query string.
extract_web: Trích xuất thông tin từ web (input: url) Input MUST be a single URL.
</tools>
<rule>
- Mỗi lần chạy chỉ gọi một tool
- search_web.input must be a search query.
- extract_web.input must be exactly one URL.
- Never output explanations inside input.
- Never output multiple URLs in one action.
- If multiple URLs are needed, create multiple actions.
- Never repeat a failed action.
- Never call the same tool with the same input twice.
- Không sử dụng các tài liệu pdf, ... để lấy dữ liệu
- Never visit URLs already listed in visited_urls.
</rule>
<guide>
Chọn ra triển lãm phù hợp để tìm các hãng dựa trên đó 
Khi có được các hãng thì cần kiểm tra xem hãng đó có đáp ứng yêu cầu không 
Sau đó là trả lời người dùng
</guide>
"""
context = ""
action_history = ""
visited_urls = set()
list_imex = []
for step in range(MAX_STEPS):

    print(f"\n===== STEP {step+1} =====")

    res = call_chat_api(messages=[{"role": "system", "content": system_prompt}, 
                        {"role": "user", "content": f"""Hãy phân tích: 
                        <list_exhibition>{data}</list_exhibition>
                        <context>{context[:50000]}</context> 
                        <history>{action_history}</history>
                        <step>{step}/{MAX_STEPS}</step>
                        <visited_urls>
                        {", ".join(visited_urls)}
                        </visited_urls>
                        <task>Hãy tìm 1-2 hãng liên quan đến sản phẩm `{query}` mà người dùng quan tâm.</task>
                        Hãy đưa ra các hành động tiếp theo để tôi có thể tìm đc các hãng theo yêu cầu dựa trên các triển lãm cho yêu cầu của người dùng.
                        Để kết thúc "next_action"="answer"
                        You are an agent, not a planner.
                        Return EXACTLY ONE action.
                        Do not plan multiple future steps.
                        Choose only the next best action based on current information.
                        Nếu số bước đạt tối đa hãy đưa ra câu trả lời cuối cùng "next_action"="answer" để kết thúc
                        Dạng trả về phải là JSON array, mỗi phần tử có dạng:
                            {{
                                "reason": "Lý do ngắn gọn (dưới 20 từ) giải thích vì sao",
                                "next_action": "hành động tiếp theo để tìm đc các hãng tham gia triển lãm cho yêu cầu của người dùng",
                                "input": "Dữ liệu đầu vào cần thiết cho hành động tiếp theo text/url"
                                "context": "Nội dung đã trích xuất từ hành động này tóm tắt lại để có thể tận dụng lại."
                                "imex": [Danh sách hãng tìm được liên quan]
                            }}
                        """
                        }
                        ], stream=False
                        )
    print(res.json().get('usage', res.json()))
    actions = extract_json(res.json()['choices'][0]['message']['content'])
    print(actions)
    action_history += f"Step: {step}/{MAX_STEPS}"
    if isinstance(actions, dict):
        actions = [actions]
    for act in actions:
        print("Nội dung", act)
        if act['next_action'] == "answer":
            list_imex = act["imex"]
            break
        data = tools[act['next_action']](act['input'])
        # print(data)
        context+=data
        if act['next_action'] == "extract_web":
            visited_urls.add(act['input'])
        action_history += f"""act: {act["next_action"]}, input: {act["input"]}"""
        list_imex = act["imex"]

# Tìm trang chủ của nhà cung cấp
for imex in list_imex:
    print(find_homepage(imex))