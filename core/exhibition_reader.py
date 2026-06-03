from dataclasses import dataclass
import openpyxl


@dataclass
class Exhibition:
    id: str
    name: str
    website: str
    category: str
    tags: str


def read_priority_exhibitions(filepath: str) -> list[Exhibition]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb["AIPT Priority Shortlist"]
    exhibitions = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 5).value
        if not name:
            continue
        exhibitions.append(Exhibition(
            id=str(ws.cell(row, 1).value or ""),
            name=str(name).strip(),
            website=str(ws.cell(row, 6).value or "").strip(),
            category=str(ws.cell(row, 7).value or "").strip(),
            tags=str(ws.cell(row, 8).value or "").strip(),
        ))
    return exhibitions
